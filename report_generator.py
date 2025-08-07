"""
Módulo de generación de reportes para la herramienta de automatización.

Este módulo se encarga de crear reportes en Excel y PDF con gráficos,
tablas resumen y análisis automático de los datos procesados.
"""

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import os
import logging
from typing import Dict, List, Tuple, Optional
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import base64

# Configurar matplotlib para generar gráficos sin interfaz gráfica
plt.switch_backend('Agg')

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Clase principal para la generación de reportes.
    
    Esta clase maneja la creación de reportes en Excel y PDF,
    incluyendo la generación automática de gráficos y tablas resumen.
    """
    
    def __init__(self):
        """Inicializa el generador de reportes."""
        self.data = None
        self.report_info = {}
        self.charts_created = []
        
        # Configurar estilo de gráficos
        plt.style.use('default')
        sns.set_palette("husl")
        
    def set_data(self, data: pd.DataFrame):
        """
        Establece los datos para el reporte.
        
        Args:
            data: DataFrame con los datos limpios
        """
        self.data = data.copy()
        logger.info(f"Datos establecidos para reporte: {len(data)} filas")
    
    def set_report_info(self, title: str, description: str = "", recipients: List[str] = None):
        """
        Establece la información del reporte.
        
        Args:
            title: Título del reporte
            description: Descripción opcional
            recipients: Lista de destinatarios
        """
        self.report_info = {
            'title': title,
            'description': description,
            'recipients': recipients or [],
            'generated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'data_rows': len(self.data) if self.data is not None else 0
        }
        logger.info(f"Información del reporte establecida: {title}")
    
    def generate_excel_report(self, output_path: str) -> Tuple[bool, str]:
        """
        Genera un reporte completo en Excel con gráficos y tablas.
        
        Esta función crea un archivo Excel con múltiples hojas:
        - Resumen ejecutivo
        - Datos originales
        - Gráficos automáticos
        - Análisis estadístico
        
        Args:
            output_path: Ruta donde guardar el archivo Excel
            
        Returns:
            Tuple con (éxito, mensaje)
        """
        if self.data is None or self.data.empty:
            return False, "No hay datos para generar el reporte."
        
        try:
            # Crear un nuevo libro de trabajo
            wb = Workbook()
            
            # Crear hojas del reporte
            self._create_summary_sheet(wb)
            self._create_data_sheet(wb)
            self._create_charts_sheet(wb)
            self._create_analysis_sheet(wb)
            
            # Guardar el archivo
            wb.save(output_path)
            logger.info(f"Reporte Excel generado exitosamente: {output_path}")
            
            return True, f"Reporte Excel generado: {output_path}"
            
        except Exception as e:
            logger.error(f"Error al generar reporte Excel: {str(e)}")
            return False, f"Error al generar reporte Excel: {str(e)}"
    
    def _create_summary_sheet(self, wb):
        """
        Crea la hoja de resumen ejecutivo.
        
        Esta hoja contiene la información general del reporte,
        estadísticas básicas y un resumen de los hallazgos principales.
        """
        ws = wb.active
        ws.title = "Resumen Ejecutivo"
        
        # Título del reporte
        ws['A1'] = self.report_info.get('title', 'Reporte Automático')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Información del reporte
        ws['A3'] = "Información del Reporte"
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = f"Generado el: {self.report_info.get('generated_at', 'N/A')}"
        ws['A5'] = f"Total de registros: {self.report_info.get('data_rows', 0):,}"
        ws['A6'] = f"Columnas analizadas: {len(self.data.columns)}"
        
        # Estadísticas básicas
        ws['A8'] = "Estadísticas Básicas"
        ws['A8'].font = Font(bold=True)
        
        row = 9
        numeric_cols = self._get_numeric_columns()
        
        for col in numeric_cols[:5]:  # Mostrar solo las primeras 5 columnas numéricas
            ws[f'A{row}'] = f"{col}:"
            ws[f'B{row}'] = f"Total: {self.data[col].sum():,.2f}"
            ws[f'C{row}'] = f"Promedio: {self.data[col].mean():,.2f}"
            ws[f'D{row}'] = f"Máximo: {self.data[col].max():,.2f}"
            row += 1
        
        # Ajustar ancho de columnas
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_data_sheet(self, wb):
        """
        Crea la hoja con los datos originales.
        
        Esta hoja contiene todos los datos procesados en formato tabla,
        con formato mejorado para facilitar la lectura.
        """
        ws = wb.create_sheet("Datos")
        
        # Agregar encabezados
        for col_num, col_name in enumerate(self.data.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Agregar datos
        for row_num, row_data in enumerate(self.data.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_charts_sheet(self, wb):
        """
        Crea la hoja con gráficos automáticos.
        
        Esta función genera gráficos automáticamente basándose en el tipo
        de datos disponible. Una imagen vale más que mil palabras cuando
        presentamos datos al jefe 😊
        """
        ws = wb.create_sheet("Gráficos")
        
        # Crear gráficos automáticos
        charts_created = []
        
        # Gráfico de barras para columnas numéricas
        numeric_cols = self._get_numeric_columns()
        if len(numeric_cols) > 0:
            chart_data = self._prepare_chart_data(numeric_cols[0])
            if chart_data is not None:
                self._add_bar_chart(ws, chart_data, f"Análisis de {numeric_cols[0]}")
                charts_created.append(f"Gráfico de barras: {numeric_cols[0]}")
        
        # Gráfico de pastel para categorías
        categorical_cols = self._get_categorical_columns()
        if len(categorical_cols) > 0 and len(numeric_cols) > 0:
            pie_data = self._prepare_pie_chart_data(categorical_cols[0], numeric_cols[0])
            if pie_data is not None:
                self._add_pie_chart(ws, pie_data, f"Distribución por {categorical_cols[0]}")
                charts_created.append(f"Gráfico de pastel: {categorical_cols[0]}")
        
        # Gráfico de línea para series temporales
        date_cols = self._get_date_columns()
        if len(date_cols) > 0 and len(numeric_cols) > 0:
            line_data = self._prepare_line_chart_data(date_cols[0], numeric_cols[0])
            if line_data is not None:
                self._add_line_chart(ws, line_data, f"Evolución temporal de {numeric_cols[0]}")
                charts_created.append(f"Gráfico de línea: {date_cols[0]}")
        
        # Información sobre gráficos creados
        ws['A1'] = "Gráficos Generados Automáticamente"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        for chart_info in charts_created:
            ws[f'A{row}'] = f"✓ {chart_info}"
            row += 1
        
        self.charts_created = charts_created
    
    def _create_analysis_sheet(self, wb):
        """
        Crea la hoja de análisis estadístico.
        
        Esta hoja contiene análisis estadísticos detallados,
        correlaciones y insights importantes de los datos.
        """
        ws = wb.create_sheet("Análisis")
        
        # Título
        ws['A1'] = "Análisis Estadístico Detallado"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Estadísticas descriptivas
        ws['A3'] = "Estadísticas Descriptivas"
        ws['A3'].font = Font(bold=True)
        
        row = 4
        numeric_cols = self._get_numeric_columns()
        
        for col in numeric_cols:
            ws[f'A{row}'] = f"Columna: {col}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            stats = self.data[col].describe()
            ws[f'B{row}'] = "Conteo:"
            ws[f'C{row}'] = f"{stats['count']:,.0f}"
            row += 1
            
            ws[f'B{row}'] = "Promedio:"
            ws[f'C{row}'] = f"{stats['mean']:,.2f}"
            row += 1
            
            ws[f'B{row}'] = "Desv. Estándar:"
            ws[f'C{row}'] = f"{stats['std']:,.2f}"
            row += 1
            
            ws[f'B{row}'] = "Mínimo:"
            ws[f'C{row}'] = f"{stats['min']:,.2f}"
            row += 1
            
            ws[f'B{row}'] = "Máximo:"
            ws[f'C{row}'] = f"{stats['max']:,.2f}"
            row += 2
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _get_numeric_columns(self) -> List[str]:
        """Obtiene las columnas numéricas del dataset."""
        numeric_cols = []
        for col in self.data.columns:
            if self.data[col].dtype in ['int64', 'float64', 'Int64']:
                numeric_cols.append(col)
        return numeric_cols
    
    def _get_categorical_columns(self) -> List[str]:
        """Obtiene las columnas categóricas del dataset."""
        categorical_cols = []
        for col in self.data.columns:
            if self.data[col].dtype == 'object' and col not in self._get_date_columns():
                categorical_cols.append(col)
        return categorical_cols
    
    def _get_date_columns(self) -> List[str]:
        """Obtiene las columnas de fecha del dataset."""
        date_cols = []
        for col in self.data.columns:
            if self.data[col].dtype == 'datetime64[ns]':
                date_cols.append(col)
        return date_cols
    
    def _prepare_chart_data(self, column_name: str) -> Optional[pd.DataFrame]:
        """
        Prepara datos para gráficos de barras.
        
        Args:
            column_name: Nombre de la columna a analizar
            
        Returns:
            DataFrame con datos preparados para el gráfico
        """
        try:
            # Para columnas numéricas, crear grupos por rangos
            if self.data[column_name].dtype in ['int64', 'float64', 'Int64']:
                # Crear bins para agrupar los datos
                bins = pd.cut(self.data[column_name], bins=10, include_lowest=True)
                chart_data = self.data.groupby(bins)[column_name].agg(['count', 'sum', 'mean']).reset_index()
                chart_data.columns = ['Rango', 'Cantidad', 'Total', 'Promedio']
                return chart_data
            return None
        except Exception as e:
            logger.warning(f"No se pudieron preparar datos para gráfico de {column_name}: {e}")
            return None
    
    def _prepare_pie_chart_data(self, category_col: str, value_col: str) -> Optional[pd.DataFrame]:
        """
        Prepara datos para gráficos de pastel.
        
        Args:
            category_col: Columna de categorías
            value_col: Columna de valores
            
        Returns:
            DataFrame con datos preparados para el gráfico de pastel
        """
        try:
            # Agrupar por categoría y sumar valores
            pie_data = self.data.groupby(category_col)[value_col].sum().reset_index()
            pie_data.columns = ['Categoría', 'Total']
            return pie_data.head(10)  # Limitar a top 10 para legibilidad
        except Exception as e:
            logger.warning(f"No se pudieron preparar datos para gráfico de pastel: {e}")
            return None
    
    def _prepare_line_chart_data(self, date_col: str, value_col: str) -> Optional[pd.DataFrame]:
        """
        Prepara datos para gráficos de línea temporal.
        
        Args:
            date_col: Columna de fechas
            value_col: Columna de valores
            
        Returns:
            DataFrame con datos preparados para el gráfico de línea
        """
        try:
            # Agrupar por fecha y sumar valores
            line_data = self.data.groupby(date_col)[value_col].sum().reset_index()
            line_data.columns = ['Fecha', 'Total']
            line_data = line_data.sort_values('Fecha')
            return line_data
        except Exception as e:
            logger.warning(f"No se pudieron preparar datos para gráfico de línea: {e}")
            return None
    
    def _add_bar_chart(self, ws, data: pd.DataFrame, title: str):
        """
        Agrega un gráfico de barras a la hoja.
        
        Args:
            ws: Hoja de trabajo
            data: Datos para el gráfico
            title: Título del gráfico
        """
        try:
            # Agregar datos a la hoja
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            
            # Crear gráfico
            chart = BarChart()
            chart.title = title
            chart.x_axis.title = "Categorías"
            chart.y_axis.title = "Valores"
            
            # Referencias para el gráfico
            data = Reference(ws, min_col=2, min_row=1, max_row=len(data)+1, max_col=len(data.columns))
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(data)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Agregar gráfico a la hoja
            ws.add_chart(chart, "A10")
            
        except Exception as e:
            logger.warning(f"Error al crear gráfico de barras: {e}")
    
    def _add_pie_chart(self, ws, data: pd.DataFrame, title: str):
        """
        Agrega un gráfico de pastel a la hoja.
        
        Args:
            ws: Hoja de trabajo
            data: Datos para el gráfico
            title: Título del gráfico
        """
        try:
            # Agregar datos a la hoja
            start_row = ws.max_row + 3
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            
            # Crear gráfico
            chart = PieChart()
            chart.title = title
            
            # Referencias para el gráfico
            data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(data), max_col=2)
            cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(data))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            
            # Agregar gráfico a la hoja
            ws.add_chart(chart, f"A{start_row + len(data) + 2}")
            
        except Exception as e:
            logger.warning(f"Error al crear gráfico de pastel: {e}")
    
    def _add_line_chart(self, ws, data: pd.DataFrame, title: str):
        """
        Agrega un gráfico de línea a la hoja.
        
        Args:
            ws: Hoja de trabajo
            data: Datos para el gráfico
            title: Título del gráfico
        """
        try:
            # Agregar datos a la hoja
            start_row = ws.max_row + 3
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            
            # Crear gráfico
            chart = LineChart()
            chart.title = title
            chart.x_axis.title = "Fecha"
            chart.y_axis.title = "Valores"
            
            # Referencias para el gráfico
            data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(data), max_col=2)
            cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(data))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            
            # Agregar gráfico a la hoja
            ws.add_chart(chart, f"A{start_row + len(data) + 2}")
            
        except Exception as e:
            logger.warning(f"Error al crear gráfico de línea: {e}")
    
    def generate_charts_images(self) -> List[Dict]:
        """
        Genera imágenes de gráficos para incluir en reportes PDF.
        
        Returns:
            Lista de diccionarios con información de los gráficos generados
        """
        if self.data is None or self.data.empty:
            return []
        
        charts = []
        
        try:
            # Configurar el estilo de los gráficos
            plt.style.use('default')
            sns.set_palette("husl")
            
            # Gráfico 1: Distribución de valores numéricos
            numeric_cols = self._get_numeric_columns()
            if len(numeric_cols) > 0:
                fig, ax = plt.subplots(figsize=(10, 6))
                self.data[numeric_cols[0]].hist(bins=20, ax=ax, alpha=0.7, color='skyblue', edgecolor='black')
                ax.set_title(f'Distribución de {numeric_cols[0]}', fontsize=14, fontweight='bold')
                ax.set_xlabel(numeric_cols[0])
                ax.set_ylabel('Frecuencia')
                ax.grid(True, alpha=0.3)
                
                # Guardar gráfico como imagen base64
                img_buffer = io.BytesIO()
                plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                img_buffer.seek(0)
                img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
                plt.close()
                
                charts.append({
                    'title': f'Distribución de {numeric_cols[0]}',
                    'type': 'histogram',
                    'image': img_base64
                })
            
            # Gráfico 2: Gráfico de barras para categorías
            categorical_cols = self._get_categorical_columns()
            if len(categorical_cols) > 0 and len(numeric_cols) > 0:
                fig, ax = plt.subplots(figsize=(12, 6))
                top_categories = self.data[categorical_cols[0]].value_counts().head(10)
                top_categories.plot(kind='bar', ax=ax, color='lightcoral')
                ax.set_title(f'Top 10 - {categorical_cols[0]}', fontsize=14, fontweight='bold')
                ax.set_xlabel(categorical_cols[0])
                ax.set_ylabel('Cantidad')
                plt.xticks(rotation=45, ha='right')
                ax.grid(True, alpha=0.3)
                
                # Guardar gráfico como imagen base64
                img_buffer = io.BytesIO()
                plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                img_buffer.seek(0)
                img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
                plt.close()
                
                charts.append({
                    'title': f'Top 10 - {categorical_cols[0]}',
                    'type': 'bar',
                    'image': img_base64
                })
            
            # Gráfico 3: Correlación entre variables numéricas
            if len(numeric_cols) > 1:
                fig, ax = plt.subplots(figsize=(8, 6))
                correlation_matrix = self.data[numeric_cols].corr()
                sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0, ax=ax)
                ax.set_title('Matriz de Correlación', fontsize=14, fontweight='bold')
                
                # Guardar gráfico como imagen base64
                img_buffer = io.BytesIO()
                plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                img_buffer.seek(0)
                img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Matriz de Correlación',
                    'type': 'heatmap',
                    'image': img_base64
                })
            
        except Exception as e:
            logger.error(f"Error al generar gráficos: {e}")
        
        return charts
    
    def get_report_summary(self) -> Dict:
        """
        Genera un resumen del reporte creado.
        
        Returns:
            Diccionario con información del reporte
        """
        return {
            'title': self.report_info.get('title', 'Reporte Automático'),
            'generated_at': self.report_info.get('generated_at', 'N/A'),
            'data_rows': self.report_info.get('data_rows', 0),
            'charts_created': len(self.charts_created),
            'charts_list': self.charts_created
        }
